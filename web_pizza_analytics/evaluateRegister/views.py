from django.shortcuts import render, get_object_or_404, redirect
import openpyxl.utils
from pizzaRegister.models import Pizzas, Ingredients, PizzaComposition
from .models import Evaluation, PizzeriaLocation, IngredientEvaluation, Notification
from accounts.models import CustomUser
from django.contrib.auth.decorators import login_required
import openpyxl
from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from django.db.models import Count, Avg
from django.db.models.functions import TruncDate
import uuid
import cv2
import json
from datetime import datetime


# Список всех оценок с фильтрами
@login_required(login_url='/login/')
def evaluation_list(request):
    unread_notifications_count = 0
    if request.user.is_authenticated:
        unread_notifications_count = request.user.notifications.filter(is_read=False).count()
    evaluations = Evaluation.objects.all()
    pizzas = Pizzas.objects.all()
    locations = PizzeriaLocation.objects.all()
    is_manager = request.user.is_authenticated and hasattr(request.user, 'pizzeria_location') and request.user.pizzeria_location and request.user.groups.filter(name="Менеджеры пиццерий").exists()
    pizza_id = request.GET.get('pizza')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    quality_max = request.GET.get('quality_max')
    quality_min = request.GET.get('quality_min')
    # Фильтры
    if is_manager:
        # Менеджер: фильтруем только по своей пиццерии
        evaluations = evaluations.filter(location=request.user.pizzeria_location)
        pizzas = pizzas.filter(pizzaembeddings__pizza__evaluation__location=request.user.pizzeria_location).distinct()
        locations = PizzeriaLocation.objects.filter(id=request.user.pizzeria_location.id)
    else:
        # Админ: применяем фильтры
        location_id = request.GET.get('location')

        if location_id:
            evaluations = evaluations.filter(location_id=location_id)

    if pizza_id:
        evaluations = evaluations.filter(pizza_id=pizza_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            evaluations = evaluations.filter(date__gte=start)
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            evaluations = evaluations.filter(date__lte=end)
        except ValueError:
            pass
    if quality_max or quality_min:
        try:
            evaluations = evaluations.filter(quality_percentage__range=(float(quality_min), float(quality_max)))
        except ValueError:
            pass

    # Сортировка
    sort_by = request.GET.get("sort_by", "date")
    order = request.GET.get("order", "desc")
    direction = "" if order == "asc" else "-"
    if sort_by in ["date", "quality_percentage"]:
        evaluations = evaluations.order_by(f"{direction}{sort_by}")

    context = {
        "evaluations": evaluations,
        "pizzas": pizzas,
        "locations": locations,
        "sort_by": sort_by,
        "order": order,
        "is_manager": is_manager,
        'unread_notifications_count': unread_notifications_count,
    }

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "evaluateRegister/evaluation_table_body.html", context)

    return render(request, "evaluateRegister/evaluation_list.html", context)


# Подробная страница одной оценки
@login_required(login_url='/login/')
def evaluation_detail(request, evaluation_id):
    unread_notifications_count = 0
    if request.user.is_authenticated:
        unread_notifications_count = request.user.notifications.filter(is_read=False).count()
    evaluation = get_object_or_404(Evaluation, id=evaluation_id)

    ingredient_evaluations = evaluation.ingredient_evaluations.select_related('ingredient').all()

    return render(request, 'evaluateRegister/evaluation_detail.html', {
        'evaluation': evaluation,
        'ingredient_evaluations': ingredient_evaluations,
        'unread_notifications_count': unread_notifications_count,
    })


@login_required(login_url='/login/')
def evaluation_export(request):
    evaluations = Evaluation.objects.all()

    evaluations = Evaluation.objects.all()
    pizzas = Pizzas.objects.all()
    is_manager = request.user.is_authenticated and hasattr(request.user, 'pizzeria_location') and request.user.pizzeria_location and request.user.groups.filter(name="Менеджеры пиццерий").exists()
    pizza_id = request.GET.get('pizza')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    quality_max = request.GET.get('quality_max')
    quality_min = request.GET.get('quality_min')
    # Фильтры
    if is_manager:
        # Менеджер: фильтруем только по своей пиццерии
        evaluations = evaluations.filter(location=request.user.pizzeria_location)
        pizzas = pizzas.filter(pizzaembeddings__pizza__evaluation__location=request.user.pizzeria_location).distinct()
    else:
        # Админ: применяем фильтры
        location_id = request.GET.get('location')

        if location_id:
            evaluations = evaluations.filter(location_id=location_id)

    if pizza_id:
        evaluations = evaluations.filter(pizza_id=pizza_id)
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            evaluations = evaluations.filter(date__gte=start)
        except ValueError:
            pass
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            evaluations = evaluations.filter(date__lte=end)
        except ValueError:
            pass
    if quality_max or quality_min:
        try:
            evaluations = evaluations.filter(quality_percentage__range=(float(quality_min), float(quality_max)))
        except ValueError:
            pass

    # Создание Excel файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оценки пицц"

    # Заголовки
    ws.append(["Пицца", "Адрес", "Дата оценки", "Качество (%)", "Размер корки"])

    for evaluation in evaluations:
        ws.append([
            evaluation.pizza.name,
            evaluation.location.address,
            evaluation.date.strftime("%d.%m.%Y %H:%M"),
            round(evaluation.quality_percentage, 1),
            round(evaluation.crust_size, 1)
        ])

    # Отправляем файл пользователю
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "evaluations_export.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@csrf_exempt
def create_evaluation_api(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST method allowed"}, status=405)

    # Получение данных
    token = request.POST.get('token')
    pizza_id = request.POST.get('pizza_id')
    crust_percentage = request.POST.get('crust_percentage')
    ingredients_json = request.POST.get('ingredients')
    crust_percentage = request.POST.get('crust_percentage')
    radius = request.POST.get('radius')
    shift = request.POST.get('shift')
    photo = request.FILES.get('photo')

    if not all([token, pizza_id, crust_percentage, ingredients_json, radius, shift, photo]):
        return JsonResponse({"error": "Missing required fields"}, status=400)

    # Проверка токена
    try:
        token_uuid = uuid.UUID(token)
        location = PizzeriaLocation.objects.get(token=token_uuid)
    except (PizzeriaLocation.DoesNotExist, ValueError):
        return JsonResponse({"error": "Invalid token"}, status=403)

    # Проверка пиццы
    try:
        pizza = Pizzas.objects.get(id=pizza_id)
    except Pizzas.DoesNotExist:
        return JsonResponse({"error": "Pizza not found"}, status=404)

    # Парсинг ингредиентов
    try:
        ingredients_data = json.loads(ingredients_json)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid ingredients format"}, status=400)

    if float(crust_percentage) > 100:
        return JsonResponse({"error": "Crust > 100%"}, status=400)

    # Получение эталонного процента корки из пиццы
    expected_crust_percentage = 100 * (pizza.pizza_size**2 - (pizza.pizza_size - pizza.crust_size)**2) / (pizza.pizza_size**2)  # предполагаем, что у модели Pizzas есть поле crust_percentage
    pizza_size = pizza.pizza_size
    shift = float(shift) / ((float(radius) * 2) / pizza_size)
    crust_size = pizza_size - pizza_size * (1 - (float(crust_percentage) / 100)) ** 0.5

    # Сохраняем оценку без качества пока
    evaluation = Evaluation.objects.create(
        pizza=pizza,
        location=location,
        photo=photo,
        quality_percentage=0.0,  # пересчитаем ниже
        crust_size=crust_size,
        shift=shift
    )

    # Сохраняем ингредиенты + собираем данные для расчёта
    ingredient_penalties = []

    for ingredient in ingredients_data:
        ing_name = ingredient.get('ingredient_name')
        detected_quantity = ingredient.get('detected_quantity')
        distribution = ingredient.get('distribution')

        try:
            ingredient_obj = Ingredients.objects.get(name=ing_name)

            # Сохраняем ингредиент
            IngredientEvaluation.objects.create(
                evaluation=evaluation,
                ingredient=ingredient_obj,
                detected_quantity=detected_quantity,
                distriubtion=float(distribution)
            )

            # Расчёт штрафа за ингредиенты
            ingredient_penalties.append(float(distribution))

        except Ingredients.DoesNotExist:
            continue

    # Расчёт оценки
    def calculate_quality(crust_real, crust_expected, penalties, shift, crust_size):
        crust_penalty = abs(crust_real - crust_expected) * 2
        shift_penalty = 100 * (shift / crust_size)
        print(f"Смещение: {shift_penalty},   Ошибка корки: {crust_penalty}")

        if penalties:
            avg_ingredient_penalty = 100 - sum(penalties) / len(penalties)
        else:
            avg_ingredient_penalty = 0

        print(f"crust_penalty: {avg_ingredient_penalty}")
        total_penalty = crust_penalty + avg_ingredient_penalty + shift_penalty
        quality = max(100 - total_penalty, 0)
        return round(quality, 1)

    final_quality = calculate_quality(
        crust_real=float(crust_percentage),
        crust_expected=expected_crust_percentage,
        penalties=ingredient_penalties,
        shift=shift,
        crust_size=pizza.crust_size
    )
    if final_quality < 70.0:
        # Находим всех менеджеров этой пиццерии
        managers = CustomUser.objects.filter(pizzeria_location=evaluation.location)
        for manager in managers:
            Notification.objects.create(
                user=manager,
                evaluation=evaluation,
                message=f"Новая оценка качества {final_quality:.1f}% для пиццы {evaluation.pizza.name}. Требуется внимание!"
            )

    # Обновляем качество в Evaluation
    evaluation.quality_percentage = final_quality
    evaluation.save()

    return JsonResponse({"success": True, "evaluation_id": evaluation.id, "quality": final_quality})


def gen_frames(stream_url):
    cap = cv2.VideoCapture(stream_url)
    while True:
        success, frame = cap.read()
        if not success:
            break
        _, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')


@login_required(login_url='/login/')
def stream_camera(request, location_id):
    try:
        location = PizzeriaLocation.objects.get(id=location_id)
        if not location.stream_url:
            return JsonResponse({"error": "У этой пиццерии не указана ссылка на камеру."}, status=400)
        return StreamingHttpResponse(gen_frames(location.stream_url),
                                     content_type='multipart/x-mixed-replace; boundary=frame')
    except PizzeriaLocation.DoesNotExist:
        return JsonResponse({"error": "Пиццерия не найдена"}, status=404)


@login_required(login_url='/login/')
def camera_page(request):
    unread_notifications_count = 0
    if request.user.is_authenticated:
        unread_notifications_count = request.user.notifications.filter(is_read=False).count()
    is_manager = request.user.is_authenticated and hasattr(request.user, 'pizzeria_location') and request.user.pizzeria_location and request.user.groups.filter(name="Менеджеры пиццерий").exists()

    # Фильтры
    if is_manager:
        locations = PizzeriaLocation.objects.filter(id=request.user.pizzeria_location.id)
    else:
        locations = PizzeriaLocation.objects.exclude(stream_url__isnull=True).exclude(stream_url='')
    return render(request, 'evaluateRegister/camera_page.html', {
        'locations': locations,
        'unread_notifications_count': unread_notifications_count
        })


@login_required(login_url='/login/')
def notifications_list(request):
    unread_notifications_count = 0
    if request.user.is_authenticated:
        unread_notifications_count = request.user.notifications.filter(is_read=False).count()
    notifications = request.user.notifications.all().order_by('-created_at')
    return render(request, 'evaluateRegister/notifications_list.html', {
        'notifications': notifications,
        'unread_notifications_count': unread_notifications_count
        })


@login_required(login_url='/login/')
def notification_redirect(request, notification_id):
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)

    if not notification.is_read:
        notification.is_read = True
        notification.save()

    return redirect('evaluation_detail', evaluation_id=notification.evaluation.id)


@login_required(login_url='/login/')
def check_new_notifications(request):
    new_notifications_count = request.user.notifications.filter(is_read=False).count()
    return JsonResponse({'new_count': new_notifications_count})


@login_required(login_url='/login/')
def reports_page(request):
    pizzas = Pizzas.objects.all()
    locations = PizzeriaLocation.objects.all()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    location_id = request.GET.get('location')
    pizza_id = request.GET.get('pizza')  # Новый фильтр по пицце

    evaluations = Evaluation.objects.all()

    # Менеджеру показываем только свою пиццерию
    if request.user.groups.filter(name="Менеджеры пиццерий").exists():
        evaluations = evaluations.filter(location=request.user.pizzeria_location)
        locations = PizzeriaLocation.objects.filter(id=request.user.pizzeria_location.id)
    else:
        if location_id:
            evaluations = evaluations.filter(location_id=location_id)
    if pizza_id:
        evaluations = evaluations.filter(pizza_id=pizza_id)
    if start_date:
        evaluations = evaluations.filter(date__gte=start_date)
    if end_date:
        evaluations = evaluations.filter(date__lte=end_date)

    # Процент несоответствий по типам пицц
    non_compliance_by_pizza = evaluations.values('pizza__name').annotate(
        avg_quality=Avg('quality_percentage')
    )

    # Динамика нарушений по дням
    dynamics = evaluations.annotate(date_only=TruncDate('date')).values('date_only').annotate(
        bad_count=Count('id', filter=models.Q(quality_percentage__lt=70))
    ).order_by('date_only')

    context = {
        'pizzas': pizzas,
        'locations': locations,
        'non_compliance_by_pizza': non_compliance_by_pizza,
        'dynamics': dynamics,
    }
    return render(request, 'evaluateRegister/reports_page.html', context)


@login_required(login_url='/login/')
def export_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    location_id = request.GET.get('location')
    pizza_id = request.GET.get('pizza')

    evaluations = Evaluation.objects.all()

    if request.user.groups.filter(name="Менеджеры пиццерий").exists():
        evaluations = evaluations.filter(location=request.user.pizzeria_location)
    else:
        if location_id:
            evaluations = evaluations.filter(location_id=location_id)

    if pizza_id:
        evaluations = evaluations.filter(pizza_id=pizza_id)
    if start_date:
        evaluations = evaluations.filter(date__gte=start_date)
    if end_date:
        evaluations = evaluations.filter(date__lte=end_date)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Несоответствия по пиццам"

    ws1.append(["Тип пиццы", "Среднее качество (%)"])
    for row in evaluations.values('pizza__name').annotate(avg=Avg('quality_percentage')):
        ws1.append([row['pizza__name'], round(row['avg'], 1)])

    ws2 = wb.create_sheet("Динамика нарушений")
    ws2.append(["Дата", "Число нарушений (<70%)"])
    for row in evaluations.annotate(date_only=TruncDate('date')).values('date_only').annotate(
        count=Count('id', filter=models.Q(quality_percentage__lt=70))
    ).order_by('date_only'):
        ws2.append([row['date_only'].strftime('%d.%m.%Y'), row['count']])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'
    wb.save(response)
    return response